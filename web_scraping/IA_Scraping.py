from bs4 import BeautifulSoup, SoupStrainer
import pandas as pd
import requests, os
from selenium import webdriver
from selenium.webdriver.support.ui import Select
import openpyxl
from openpyxl.styles import Font

def getPrice(string):
    price = ""
    for s in string:
        try:
            price += str(int(s))
        except:
            pass

    #insert decimal
    price = price[:-2] + "." + price[-2:]

    return "{:,.2f}".format(float(price.replace(",","")))

def getProduct(string):
    product = ""
    string = string[::-1]
    for s in string:
        if s == " ":
            break
        product += s

    return product[::-1]


base_url = """https://industrialautomationco.com"""

options = webdriver.ChromeOptions()
options.add_argument('--headless')
options.add_argument(str(os.path.join(os.path.dirname(os.path.realpath(__file__)), "chromedriver.exe")))

driver = webdriver.Chrome(options=options, executable_path=str(
    os.path.join(os.path.dirname(os.path.realpath(__file__)), "chromedriver.exe")))


#Pages 138. Each part under <a class>
beg_url = """https://industrialautomationco.com/collections/all?page=1"""

#get contents of first page
content = requests.get(beg_url).text
soup = BeautifulSoup(content, features="lxml")

#Find how many pages
text = str(soup.find(class_ = """pagination__text"""))
text = text[::-1]
total_pages = ""
##int_found: once start to find integers, once hit a space, no more need to keep going through
int_found = False
for let in text:
    try:
        int(let)
        total_pages += let
        int_found = True
    except:
        if int_found:
            break
        pass
total_pages = int(total_pages[::-1])

#go through each page and get links
page_url = """https://industrialautomationco.com/collections/all?page="""
product_links = []
for i in range(1,total_pages+1):
    content = requests.get(page_url + str(i)).text
    soup = BeautifulSoup(content, features="lxml")

    for product in soup.find_all(class_ = """grid-view-item__link grid-view-item__image-container full-width-link"""):
        product_links.append(product['href'])


#Go through each url and get prices/stocks
products = []
prices = []
types = []
stocks = []
i = 0
for prod in product_links:
    try:
        prices.append([])
        types.append([])
        stocks.append([])

        url = base_url + prod
        #print(url)
        driver.get(url)

        #Get Part Number
        product = getProduct(str(driver.find_element_by_class_name("product-single__title").text))
        products.append(product)

        # Get all types - refurbished, New no Box, New, etc.
        select_box = driver.find_element_by_id("SingleOptionSelector-0")
        options_html = [x for x in select_box.find_elements_by_tag_name("option")]
        options = []
        for el in options_html:
            options.append(el.get_attribute("value"))

        # Gets prices
        select = Select(driver.find_element_by_id("SingleOptionSelector-0"))  # I believe this is the correct id selector
        #Have to select another label first in order for first price to show up
        select.select_by_index(0)

        for el in options:
            select.select_by_visible_text(el)

            element = driver.find_element_by_class_name("price__regular")
            text = element.get_attribute('innerHTML')
            price = getPrice(text)

            prices[i].append(price)
            types[i].append(el)
            ##FOR NOW, until know how get stocks #TODO: Figure out how get Stocks
            #Look for available in the soup??
            stocks[i].append(1)
        i += 1
    except Exception as e:
        i+=1
        #print(e)

driver.close()
print(products)
print(prices)
print(types)

def get_column_names(ws):
    column_names = []
    for i in range(1, ws.max_column + 1):
        value = ws[f"{chr(i + 65)}1"].value
        if value is not None:
            column_names.append(value)
        i += 1

    return column_names

def add_item(part,price,type,ws):
    column_names = get_column_names(ws)
    if type not in column_names:
        ws[f"{chr(len(ws['1']) + 64)}1"] = type
        ws[f"{chr(len(ws['1']) + 64)}1"].font = Font(bold=True)

    for j in range(1,ws.max_column+1):
        if ws[f"{chr(64 + j)}1"].value == type:
            break
    col = j

    done = False
    for i in range(1,ws.max_row + 1):
        if ws[f"A{i}"].value == part:
            ws[f"{chr(64+col)}{i}"] = price
            done = True
    if not done:
        ws[f"A{ws.max_row + 1}"] = part
        ws[f"{chr(64 + col)}{ws.max_row}"] = price


wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = "Part Number"
ws["A1"].font = Font(bold=True)
for i in range(len(products)):
    for j in range(len(prices[i])):
        add_item(products[i],prices[i][j],types[i][j],ws)
ws.freeze_panes = 'A2'
wb.save(str(os.path.join(os.path.dirname(os.path.realpath(__file__)),"Industrial_Automation_Prices.xlsx")))

##FOR ACCESS DATABASES
#import pyodbc
#global connection
#global cursor
def connectDatabase(database_file):
    global connection
    global cursor

    try:
        connection = pyodbc.connect(r'Driver={Microsoft Access Driver (*.mdb, *.accdb)};DBQ='+database_file+';')

        #connection = pyodbc.connect(r'DSN=MSSQL-PYTHON;DRIVER={SQL Server};DBQ='+database_file+';')
        cursor = connection.cursor()
    except Exception as e:
        try:
            print(e)
            connection = pyodbc.connect(r"Driver={SQL Server};DBQ=" + database_file + ';')
            cursor = connection.cursor()
        except Exception as e:
            print(e)
            try:
                connection = pyodbc.connect(r"Driver={SQL Server};server='MS Access Database';DBQ=" + database_file + ';')
                cursor = connection.cursor()
            except Exception as e:
                print(e)
                connection = pyodbc.connect(r"Driver={SQL Server};server='MS Access Database';DBQ=" + database_file + ';Trusted_Connection=yes;')
                cursor = connection.cursor()

def closeDatabaseConnection():
    global connection
    connection.close()
IA_table = "Industrial_Auto"
def del_table_ACCESS():
    global cursor
    cursor.execute(f"select * from {IA_table}")
    column_names = [i[0] for i in cursor.description]
    i = 0
    for col in column_names:
        if i == 0:
            #Don't delete main column
            i += 1
        else:
            cursor.execute(f"ALTER TABLE {IA_table} DROP {col}")

    #Delete all rows in table
    cursor.execute(f"DELETE FROM {IA_table}")

def add_item_ACCESS(part,price,type):
    global cursor
    global connection
    cursor.execute(f"select * from {IA_table}")
    column_names = [i[0] for i in cursor.description]
    if type not in column_names:
        cursor.execute(f"ALTER TABLE {IA_table} ADD {type} CURRENCY")
    cursor.execute(f"INSERT into {IA_table} (Part_Number,{type}) values ('{part}','{price}')")


